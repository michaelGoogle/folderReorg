"""
Phase 1b — extract the first ~4000 characters of text from each file (plan §5.5).

Runs in a process pool; writes one .txt per file keyed by file_id, plus an
extraction_results.csv with (file_id, status, text_chars, text_path).

Status values:
    ok                         - extracted successfully
    quarantine_image_only      - PDF with no extractable text
    quarantine_password        - encrypted PDF
    quarantine_corrupt         - read error / parser failure
    quarantine_too_large       - > MAX_SIZE
    quarantine_no_extractor    - binary/image/archive (handled by non-text path in phase 3)

Non-text extensions are NOT an error — they are flagged `quarantine_no_extractor`
and handled by the parent-folder rule in phase 3 (§7.6, §11.1).
"""

from __future__ import annotations

import argparse
import csv
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from tqdm import tqdm

from src.config import DATA_DIR, EXTRACTED_TEXT_DIR, EXTRACT_WORKERS, MAX_TEXT_CHARS

MAX_SIZE = 500 * 1024 * 1024   # 500 MB — §11.7

# --- per-format extractors -------------------------------------------------


def _extract_pdf(path: Path) -> tuple[str, str]:
    try:
        import fitz  # PyMuPDF (imported lazily — heavy)
        with fitz.open(path) as doc:
            if doc.is_encrypted:
                return "quarantine_password", ""
            text = ""
            for page in doc:
                text += page.get_text("text")
                if len(text) >= MAX_TEXT_CHARS:
                    break
            text = text.strip()
            if len(text) < 50:  # heuristic: effectively no extractable text
                return "quarantine_image_only", ""
            return "ok", text[:MAX_TEXT_CHARS]
    except Exception:
        return "quarantine_corrupt", ""


def _extract_docx(path: Path) -> tuple[str, str]:
    try:
        from docx import Document as Docx
        doc = Docx(path)
        text = "\n".join(p.text for p in doc.paragraphs if p.text and p.text.strip())
        return ("ok", text[:MAX_TEXT_CHARS]) if text else ("quarantine_corrupt", "")
    except Exception:
        return "quarantine_corrupt", ""


def _extract_xlsx(path: Path) -> tuple[str, str]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        for ws in wb.worksheets[:3]:
            parts.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(max_row=20, values_only=True):
                parts.append("\t".join("" if v is None else str(v) for v in row))
        return "ok", "\n".join(parts)[:MAX_TEXT_CHARS]
    except Exception:
        return "quarantine_corrupt", ""


def _extract_text(path: Path) -> tuple[str, str]:
    try:
        return "ok", path.read_text(errors="ignore")[:MAX_TEXT_CHARS]
    except Exception:
        return "quarantine_corrupt", ""


EXTRACTORS = {
    ".pdf":  _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
    ".xlsm": _extract_xlsx,
    ".txt":  _extract_text,
    ".md":   _extract_text,
    ".csv":  _extract_text,
    # .doc / .xls (old binary formats) deliberately not included — will quarantine
}


def _process_one(row: dict) -> tuple[str, str, int, str]:
    """Runs in a worker process. Returns (file_id, status, text_chars, text_path)."""
    path = Path(row["abs_path"])
    try:
        size = int(row["size_bytes"])
    except (KeyError, TypeError, ValueError):
        size = path.stat().st_size if path.exists() else 0
    if size > MAX_SIZE:
        return row["file_id"], "quarantine_too_large", 0, ""

    extractor = EXTRACTORS.get(row["ext"])
    if extractor is None:
        # Images, videos, archives, .doc, .xls, etc. → handled by non-text path.
        return row["file_id"], "quarantine_no_extractor", 0, ""

    status, text = extractor(path)
    if status == "ok":
        out = EXTRACTED_TEXT_DIR / f"{row['file_id']}.txt"
        out.write_text(text, encoding="utf-8")
        return row["file_id"], status, len(text), str(out)
    return row["file_id"], status, 0, ""


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--inventory", type=Path, default=DATA_DIR / "inventory.csv")
    ap.add_argument("--out",       type=Path, default=DATA_DIR / "extraction_results.csv")
    ap.add_argument("--workers",   type=int,  default=EXTRACT_WORKERS)
    args = ap.parse_args()

    inv = pd.read_csv(args.inventory)
    records = inv.to_dict("records")
    results: list[tuple[str, str, int, str]] = []
    with ProcessPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(_process_one, r): r for r in records}
        for fut in tqdm(as_completed(futures), total=len(futures), unit="file", desc="extract"):
            results.append(fut.result())

    with args.out.open("w", newline="", encoding="utf-8") as w:
        writer = csv.writer(w)
        writer.writerow(["file_id", "status", "text_chars", "text_path"])
        writer.writerows(results)
    # Summary
    df = pd.DataFrame(results, columns=["file_id", "status", "text_chars", "text_path"])
    print(df["status"].value_counts().to_string())
    print(f"OK — wrote {len(results):,} rows to {args.out}")


if __name__ == "__main__":
    main()
